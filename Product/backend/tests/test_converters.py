"""Tests for converters."""

import json
import tempfile
from pathlib import Path

from backend.converters.json_to_dmn import convert_to_dmn
from backend.converters.json_to_mermaid import convert_to_mermaid
from backend.converters.json_to_csv import convert_to_csv, convert_predicates_to_csv, convert_phrases_to_csv
from backend.converters.json_to_xlsx import convert_to_xlsx

FIXTURES = Path(__file__).parent / "fixtures"


def load_fixture(name: str) -> dict:
    with open(FIXTURES / name) as f:
        return json.load(f)


class TestDMNConverter:
    def test_produces_valid_xml(self):
        logic = load_fixture("valid_logic.json")
        xml = convert_to_dmn(logic)
        assert "<?xml" in xml
        assert "definitions" in xml

    def test_contains_decision_tables(self):
        logic = load_fixture("valid_logic.json")
        xml = convert_to_dmn(logic)
        assert "decisionTable" in xml
        assert "hitPolicy" in xml

    def test_contains_module_decisions(self):
        logic = load_fixture("valid_logic.json")
        xml = convert_to_dmn(logic)
        assert "mod_cough" in xml

    def test_contains_provenance(self):
        logic = load_fixture("valid_logic.json")
        xml = convert_to_dmn(logic)
        assert "p.38" in xml or "p.22" in xml

    def test_contains_activator(self):
        logic = load_fixture("valid_logic.json")
        xml = convert_to_dmn(logic)
        assert "COLLECT" in xml


class TestMermaidConverter:
    def test_produces_flowchart(self):
        logic = load_fixture("valid_logic.json")
        mermaid = convert_to_mermaid(logic)
        assert "graph TD" in mermaid

    def test_contains_modules(self):
        logic = load_fixture("valid_logic.json")
        mermaid = convert_to_mermaid(logic)
        assert "mod_cough" in mermaid

    def test_contains_emergency_path(self):
        logic = load_fixture("valid_logic.json")
        mermaid = convert_to_mermaid(logic)
        assert "emergency" in mermaid.lower()


class TestCSVConverter:
    def test_predicates_csv_has_headers(self):
        logic = load_fixture("valid_logic.json")
        csv_str = convert_predicates_to_csv(logic)
        assert "predicate_id" in csv_str
        assert "threshold_expression" in csv_str

    def test_predicates_csv_has_data(self):
        logic = load_fixture("valid_logic.json")
        csv_str = convert_predicates_to_csv(logic)
        assert "p_danger_sign_present" in csv_str
        assert "p_fast_breathing" in csv_str

    def test_phrases_csv_has_headers(self):
        logic = load_fixture("valid_logic.json")
        csv_str = convert_phrases_to_csv(logic)
        assert "message_id" in csv_str
        assert "english_text" in csv_str

    def test_phrases_csv_has_data(self):
        logic = load_fixture("valid_logic.json")
        csv_str = convert_phrases_to_csv(logic)
        assert "m_dx_pneumonia" in csv_str

    def test_convert_to_csv_returns_both(self):
        logic = load_fixture("valid_logic.json")
        result = convert_to_csv(logic)
        assert "predicates" in result
        assert "phrases" in result


class TestXLSXConverter:
    def test_creates_file(self):
        logic = load_fixture("valid_logic.json")
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = convert_to_xlsx(logic, f.name)
            assert Path(path).exists()
            assert Path(path).stat().st_size > 0

    def test_has_required_sheets(self):
        from openpyxl import load_workbook
        logic = load_fixture("valid_logic.json")
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            convert_to_xlsx(logic, f.name)
            wb = load_workbook(f.name)
            assert "survey" in wb.sheetnames
            assert "choices" in wb.sheetnames
            assert "settings" in wb.sheetnames
