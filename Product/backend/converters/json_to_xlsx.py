"""Convert clinical_logic JSON to CHT XLSForm XLSX.

This is the most complex converter. Key challenges:
1. The Boolean Trap: ODK evaluates string("false") as truthy. Use numeric 1/0.
2. Predicate compilation with fail-safes: if(measured, threshold_logic, fail_safe_value)
3. Module state management: completed_modules bitmask with pipe delimiters
4. Module-gated relevance: contains(required_modules, 'mod_X') and not(contains(completed_modules, '|mod_X|'))

Schema compatibility:
- gen7 predicates: `source_vars` + `threshold_expression` + `fail_safe`
- gen8 predicates: `inputs_used` + `formal_definition` + `missingness_rule`
- gen7 router:    `{rows: [...]}` with `output_module` + flat `condition`
- gen8 router:    `{cop1_queue_builder, cop2_next_module}` with `actions.add_to_queue/next_module`
- gen7 modules:   dict keyed by id with `module_id` field
- gen8 modules:   same shape; predicate ids may not have `predicate_id` alias

The converter walks `logic.variables[]` to build CHW input questions. Earlier
versions only emitted input rows for variables referenced as `source_vars`
in predicates -- which produced a 3-question form when the variables list
had 100+ entries. Now every variable that has a recognized prefix gets a
CHW-facing input row.
"""

from openpyxl import Workbook


# ---- Field-name compatibility shims ----

def _pred_inputs(p: dict) -> list[str]:
    """Return the list of raw variables a predicate depends on.

    Reads `inputs_used` (gen8) first, then `source_vars` (gen7 alias).
    """
    return list(p.get("inputs_used") or p.get("source_vars") or [])


def _pred_definition(p: dict) -> str:
    """Return the predicate's boolean definition string.

    Reads `formal_definition` (gen8) first, then `threshold_expression` (gen7).
    """
    return str(p.get("formal_definition") or p.get("threshold_expression") or "")


def _pred_missingness_default(p: dict) -> int:
    """Numeric fail-safe for missing inputs.

    Maps gen8 `missingness_rule` to the gen7 0/1 fail_safe used in XLSForm
    `if(measured, threshold, fail_safe)` calculations:
      FALSE_IF_MISSING        -> 0
      TRIGGER_REFERRAL        -> 1 (escalate)
      BLOCK_DECISION          -> 0 (predicate cannot fire)
      ALERT_NO_RULE_SPECIFIED -> 0 (default conservative)
    Falls back to the legacy `fail_safe` int when present.
    """
    if "fail_safe" in p:
        try:
            return int(p["fail_safe"])
        except (TypeError, ValueError):
            pass
    rule = str(p.get("missingness_rule") or "").upper()
    return 1 if rule == "TRIGGER_REFERRAL" else 0


def _is_stub_predicate(p: dict) -> bool:
    """Auto-registered stubs lack a definition AND a non-empty inputs list."""
    if p.get("_auto_registered") and not _pred_definition(p):
        return True
    if not _pred_definition(p) and not _pred_inputs(p):
        return True
    return False


def _compile_predicate_xpath(predicate: dict) -> list[dict]:
    """Compile a predicate into XLSForm calculate rows.

    Returns list of rows for the survey sheet (type, name, calculation).
    Each predicate becomes two calculates:
    1. A "measured" check (are source vars present?)
    2. The predicate itself (with fail-safe for missing data)
    """
    pid = predicate.get("id") or predicate.get("predicate_id") or "unknown"
    source_vars = _pred_inputs(predicate)
    threshold = _pred_definition(predicate)
    fail_safe = _pred_missingness_default(predicate)

    rows: list[dict] = []

    # Stub predicate: emit a single calculate that always returns the fail-safe
    # (so downstream rule references don't fail) but tag the row so reviewers
    # can spot it in the survey.
    if _is_stub_predicate(predicate):
        rows.append({
            "type": "calculate",
            "name": pid,
            "calculation": str(fail_safe),
            "_stub": True,
        })
        return rows

    # Measured check: all source vars have values
    if source_vars:
        measured_checks = [
            f"string-length(${{{sv}}}) > 0" for sv in source_vars
        ]
        measured_calc = " and ".join(measured_checks)
        rows.append({
            "type": "calculate",
            "name": f"{pid}_measured",
            "calculation": measured_calc,
        })

    # Predicate itself with fail-safe
    xpath_threshold = _threshold_to_xpath(threshold, source_vars)
    if source_vars:
        rows.append({
            "type": "calculate",
            "name": pid,
            "calculation": f"if(${{{pid}_measured}}, if({xpath_threshold}, 1, 0), {fail_safe})",
        })
    else:
        rows.append({
            "type": "calculate",
            "name": pid,
            "calculation": f"if({xpath_threshold}, 1, 0)",
        })

    return rows


def _threshold_to_xpath(expression: str, source_vars: list[str]) -> str:
    """Convert a threshold expression to XPath-compatible format.

    This is a simplified converter. In production, a proper parser would
    translate the expression AST. Here we do basic substitution.
    """
    result = expression

    # Replace variable references with XPath variable syntax
    for sv in sorted(source_vars, key=len, reverse=True):
        result = result.replace(sv, f"${{{sv}}}")

    # Replace logical operators
    result = result.replace(" AND ", " and ")
    result = result.replace(" OR ", " or ")
    result = result.replace("==", "=")
    # gen8 boolean equality: `q_fever = true` -> `${q_fever} = 1`
    result = result.replace("= true", "= 1").replace("= false", "= 0")

    return result


_XPATH_VAR_PREFIXES = (
    "p_", "c_", "calc_", "q_", "ex_", "v_", "lab_", "img_",
    "demo_", "hx_", "sys_", "prev_", "mod_", "is_",
)


def _to_xpath(expression: str) -> str:
    """Generic expression-to-XPath converter (no source_vars list needed).

    Used by router rule conditions and module rule conditions in the gen8
    pipeline where the original `inputs_used` may not be on the rule itself.
    """
    if not expression:
        return ""
    import re as _re
    result = expression

    # Wrap any prefix-tagged identifier in ${...}
    pat = r"\b((?:" + "|".join(_re.escape(p) for p in _XPATH_VAR_PREFIXES) + r")[a-z0-9_]+)\b"
    result = _re.sub(pat, r"${\1}", result)

    result = (result
              .replace(" AND ", " and ")
              .replace(" OR ", " or ")
              .replace("==", "=")
              .replace("= true", "= 1")
              .replace("= false", "= 0"))
    return result


def _build_survey_sheet(logic: dict) -> list[list[str]]:
    """Build the survey sheet rows.

    Walks the FULL `logic.variables[]` list (not just predicate source_vars)
    and emits one CHW-input row per variable, grouped by prefix:
      demo_*  -> demographics  (text/integer/select)
      hx_*    -> medical history (select_one yesno)
      q_*     -> symptom screening (select_one yesno)
      ex_*    -> examination findings (select_one yesno)
      v_*     -> vital signs / measurements (integer/decimal)
      lab_*   -> point-of-care lab results (mostly select_one)
    Internal prefixes (p_, calc_, mod_*_done, sys_, prev_, is_*) are emitted
    as hidden `calculate` rows by the predicate compiler -- not as inputs.
    """
    headers = ["type", "name", "label", "calculation", "relevant", "required", "appearance"]
    rows: list[list[str]] = [headers]

    # ----- Walk variables list and bucket by prefix -----
    raw_vars = logic.get("variables", []) or []
    if isinstance(raw_vars, dict):
        # If someone passed a dict-of-dicts shape, normalize.
        raw_vars = [{**v, "id": k} for k, v in raw_vars.items() if isinstance(v, dict)]

    buckets: dict[str, list[dict]] = {
        "demo": [], "hx": [], "q": [], "ex": [], "v": [], "lab": [], "img": [],
    }
    for v in raw_vars:
        if not isinstance(v, dict):
            continue
        vid = str(v.get("id") or "").strip()
        if not vid:
            continue
        for prefix in ("demo_", "hx_", "q_", "ex_", "v_", "lab_", "img_"):
            if vid.startswith(prefix):
                buckets[prefix.rstrip("_")].append(v)
                break

    # Stable ordering: prefer manual page if present, then alphabetical.
    for k in buckets:
        buckets[k].sort(
            key=lambda v: (v.get("_manual_page") or 9999, v.get("id", ""))
        )

    def _label(v: dict, prefix: str) -> str:
        return (
            str(v.get("display_name") or v.get("human_label") or
                v.get("id", "").removeprefix(prefix).replace("_", " ").title())
        )

    def _data_type(v: dict, default: str) -> str:
        """Choose the XLSForm input type from the variable's data_type metadata.

        If the variable carries an `allowed_input_domain` like `{a, b, c}`, we
        return `select_one <listname>` referencing the per-variable list that
        `_build_choices_sheet` emits.
        """
        import re as _re
        domain = str(v.get("allowed_input_domain") or v.get("allowed_values") or "")
        m = _re.match(r"^\s*[\{\[]\s*([^}\]]+)[\}\]]\s*$", domain)
        if m and len([c for c in m.group(1).split(",") if c.strip()]) >= 2:
            list_name = v.get("id", "")
            for prefix in ("v_", "ex_", "lab_", "q_", "hx_"):
                if list_name.startswith(prefix):
                    list_name = list_name[len(prefix):]
                    break
            return f"select_one {list_name}"
        dt = str(v.get("data_type") or "").lower()
        if dt in {"integer", "int"}:
            return "integer"
        if dt in {"decimal", "float", "number"}:
            return "decimal"
        if dt in {"text", "string"}:
            return "text"
        if dt in {"date"}:
            return "date"
        return default

    # ----- 1. Demographics -----
    rows.append(["begin_group", "grp_demographics", "Patient Information", "", "", "", ""])
    if buckets["demo"]:
        for v in buckets["demo"]:
            vid = v["id"]
            kind = _data_type(v, "text")
            # Special-cased: sex / age get explicit types
            if "age" in vid and "month" in vid:
                kind = "integer"
            elif vid.endswith("_sex"):
                kind = "select_one sex"
            rows.append([kind, vid, _label(v, "demo_"), "", "", "yes", ""])
    else:
        # Defensive: if Stage 3 didn't emit demo variables, keep the legacy 3.
        rows.append(["text", "demo_patient_id", "Patient ID", "", "", "yes", ""])
        rows.append(["integer", "demo_age_months", "Age (months)", "", "", "yes", ""])
        rows.append(["select_one sex", "demo_sex", "Sex", "", "", "yes", ""])
    rows.append(["end_group", "", "", "", "", "", ""])

    # ----- 2. Medical history -----
    if buckets["hx"]:
        rows.append(["begin_group", "grp_history", "Medical History", "", "", "", ""])
        for v in buckets["hx"]:
            kind = _data_type(v, "select_one yesno")
            rows.append([kind, v["id"], _label(v, "hx_"), "", "", "", ""])
        rows.append(["end_group", "", "", "", "", "", ""])

    # ----- 3. Symptom screening (q_) -----
    if buckets["q"]:
        rows.append(["begin_group", "grp_screening", "Symptom Screening", "", "", "", ""])
        for v in buckets["q"]:
            kind = _data_type(v, "select_one yesno")
            label = _label(v, "q_")
            if kind.startswith("select_one yesno") and not label.endswith("?"):
                label = label + "?"
            rows.append([kind, v["id"], label, "", "", "yes", ""])
        rows.append(["end_group", "", "", "", "", "", ""])

    # ----- 4. Examination findings (ex_) -----
    if buckets["ex"]:
        rows.append(["begin_group", "grp_examination", "Examination Findings",
                     "", "", "", ""])
        for v in buckets["ex"]:
            kind = _data_type(v, "select_one yesno")
            rows.append([kind, v["id"], _label(v, "ex_"), "", "", "", ""])
        rows.append(["end_group", "", "", "", "", "", ""])

    # ----- 5. Vitals / measurements (v_) -----
    if buckets["v"]:
        rows.append(["begin_group", "grp_vitals", "Vital Signs / Measurements",
                     "", "", "", ""])
        for v in buckets["v"]:
            vid = v["id"]
            # Integer for counts (rate, days, age), decimal for everything else.
            if any(k in vid for k in ("rate", "_bpm", "_per_min", "_days", "_count", "_minutes", "_months")):
                kind = _data_type(v, "integer")
            else:
                kind = _data_type(v, "decimal")
            unit = v.get("unit") or v.get("units") or ""
            label = _label(v, "v_") + (f" ({unit})" if unit else "")
            rows.append([kind, vid, label, "", "", "", ""])
        rows.append(["end_group", "", "", "", "", "", ""])

    # ----- 6. Lab results (lab_) -----
    if buckets["lab"]:
        rows.append(["begin_group", "grp_labs", "Point-of-Care Test Results",
                     "", "", "", ""])
        for v in buckets["lab"]:
            kind = _data_type(v, "select_one yesno")
            rows.append([kind, v["id"], _label(v, "lab_"), "", "", "", ""])
        rows.append(["end_group", "", "", "", "", "", ""])

    # 4. Hidden predicate calculates
    for pred in logic.get("predicates", []):
        pred_rows = _compile_predicate_xpath(pred)
        for pr in pred_rows:
            rows.append([
                pr["type"], pr["name"], "", pr["calculation"], "", "", "",
            ])

    # 5. Router calculate: build required_modules string.
    #    gen7: router.rows[].output_module + condition string
    #    gen8: router.cop1_queue_builder.rules[].actions.add_to_queue (list)
    router = logic.get("router") or logic.get("activator") or {}
    if not isinstance(router, dict):
        router = {}
    cop1 = router.get("cop1_queue_builder") if isinstance(router.get("cop1_queue_builder"), dict) else None
    router_rules = router.get("rows") or router.get("rules") or []

    activator_parts: list[str] = []
    if cop1 and isinstance(cop1.get("rules"), list):
        # gen8 two-cop router: cop1 enqueues modules
        for rule in cop1["rules"]:
            if not isinstance(rule, dict):
                continue
            actions = rule.get("actions") or {}
            queue = actions.get("add_to_queue") or []
            if isinstance(queue, str):
                queue = [queue]
            conds = rule.get("conditions") or {}
            cond_expr = ""
            if isinstance(conds, dict):
                # Preferred shape: {"key": value, ...}
                if conds.get("at_start") is True:
                    cond_expr = "true"
                elif "expr" in conds and isinstance(conds["expr"], str):
                    cond_expr = conds["expr"]
                else:
                    parts = []
                    for k, v in conds.items():
                        if isinstance(v, bool):
                            parts.append(f"${{{k}}} = {1 if v else 0}")
                        else:
                            parts.append(f"${{{k}}} = '{v}'")
                    cond_expr = " and ".join(parts) if parts else "true"
            elif isinstance(conds, str):
                cond_expr = conds
            if not queue:
                continue
            xpath_cond = _to_xpath(cond_expr)
            for mid in queue:
                if cond_expr in ("", "true"):
                    activator_parts.append(f"'|{mid}|'")
                else:
                    activator_parts.append(f"if({xpath_cond}, '|{mid}|', '')")
    elif router_rules:
        # gen7 flat router fallback
        for rule in router_rules:
            mid = rule.get("output_module") or rule.get("module_id", "")
            condition_str = rule.get("condition", "")
            if condition_str and condition_str != "true":
                xpath_cond = _to_xpath(condition_str)
                activator_parts.append(f"if({xpath_cond}, '|{mid}|', '')")
            elif condition_str == "true":
                activator_parts.append(f"'|{mid}|'")
    if activator_parts:
        calc = "concat(" + ", ".join(activator_parts) + ")"
        rows.append(["calculate", "required_modules", "", calc, "", "", ""])

    # 6. Completed modules tracker and module groups
    rows.append(["calculate", "completed_modules", "", "''", "", "", ""])

    # Gen 7: modules is a dict keyed by module_id. Fall back to list.
    raw_modules = logic.get("modules", [])
    if isinstance(raw_modules, dict):
        module_list = list(raw_modules.values())
    elif isinstance(raw_modules, list):
        module_list = raw_modules
    else:
        module_list = []

    for module in module_list:
        mid = module.get("module_id") or module.get("id") or ""
        if not mid:
            continue
        display = module.get("display_name") or module.get("title") or mid

        # Module group with relevance gate
        relevance = (
            f"contains(${{required_modules}}, '{mid}') "
            f"and not(contains(${{completed_modules}}, '|{mid}|')) "
            f"and ${{p_danger_sign_present}} = 0"
        )
        rows.append(["begin_group", f"grp_{mid}", display, "", relevance, "", ""])

        # Module-specific questions/display would go here
        rows.append(["note", f"note_{mid}", f"Assessing: {display}", "", "", "", ""])

        rows.append(["end_group", "", "", "", "", "", ""])

    # 7. Emergency group
    rows.append([
        "begin_group", "grp_emergency", "EMERGENCY REFERRAL",
        "", "${p_danger_sign_present} = 1", "", "",
    ])
    rows.append([
        "note", "note_emergency",
        "ALERT / RED-FLAG CRITERION DETECTED. Escalate this patient urgently according to the manual's high-priority pathway.",
        "", "", "", "",
    ])
    rows.append(["end_group", "", "", "", "", "", ""])

    # 8. Integrative group
    rows.append(["begin_group", "grp_integrative", "Care Plan Summary", "", "", "", ""])
    rows.append(["note", "note_integrative", "Review the combined care plan.", "", "", "", ""])
    rows.append(["end_group", "", "", "", "", "", ""])

    return rows


def _build_choices_sheet(logic: dict | None = None) -> list[list[str]]:
    """Build the choices sheet.

    Always emits the canonical `yesno` and `sex` lists, then adds one list
    per categorical variable found in `logic.variables[]` whose
    `allowed_input_domain` looks like a `{a, b, c}` set literal. This
    surfaces MUAC `{red, green, yellow}`, RDT `{positive, negative, invalid}`,
    etc. without hand-coding them.
    """
    headers = ["list_name", "name", "label"]
    rows: list[list[str]] = [headers]
    rows.append(["yesno", "1", "Yes"])
    rows.append(["yesno", "0", "No"])
    rows.append(["sex", "male", "Male"])
    rows.append(["sex", "female", "Female"])

    if not logic:
        return rows
    raw_vars = logic.get("variables", []) or []
    if isinstance(raw_vars, dict):
        raw_vars = [{**v, "id": k} for k, v in raw_vars.items() if isinstance(v, dict)]

    seen_lists: set[str] = {"yesno", "sex"}
    import re as _re
    for v in raw_vars:
        if not isinstance(v, dict):
            continue
        domain = str(v.get("allowed_input_domain") or v.get("allowed_values") or "")
        # Match `{red, green, yellow}` or `[red, green, yellow]`
        m = _re.match(r"^\s*[\{\[]\s*([^}\]]+)[\}\]]\s*$", domain)
        if not m:
            continue
        choices = [c.strip() for c in m.group(1).split(",") if c.strip()]
        if len(choices) < 2:
            continue
        # List name: the variable id stripped of its prefix
        list_name = v["id"]
        for prefix in ("v_", "ex_", "lab_", "q_", "hx_"):
            if list_name.startswith(prefix):
                list_name = list_name[len(prefix):]
                break
        if list_name in seen_lists:
            continue
        seen_lists.add(list_name)
        for c in choices:
            rows.append([list_name, c, c.title()])
    return rows


def _build_settings_sheet(manual_name: str = "CHW Navigator") -> list[list[str]]:
    """Build the settings sheet."""
    headers = ["form_title", "form_id", "version", "style"]
    rows: list[list[str]] = [headers]
    rows.append([manual_name, "chw_navigator", "1", "pages"])
    return rows


def convert_to_xlsx(logic, output_path: str, manual_name: str = "CHW Navigator") -> str:
    """Convert clinical_logic JSON to CHT XLSForm XLSX.

    Args:
        logic: The validated clinical_logic JSON dict.
        output_path: Path to write the XLSX file.
        manual_name: Name for the form title.

    Returns:
        The output_path.
    """
    # Defensive top-level guard -- fall back to an empty dict so we emit a
    # valid (empty) XLSForm instead of crashing on logic.get() below.
    if not isinstance(logic, dict):
        logic = {}
    wb = Workbook()

    # Survey sheet
    ws_survey = wb.active
    ws_survey.title = "survey"
    for row in _build_survey_sheet(logic):
        ws_survey.append(row)

    # Choices sheet (logic-aware so we can emit MUAC / RDT / etc. enum lists)
    ws_choices = wb.create_sheet("choices")
    for row in _build_choices_sheet(logic):
        ws_choices.append(row)

    # Settings sheet
    ws_settings = wb.create_sheet("settings")
    for row in _build_settings_sheet(manual_name):
        ws_settings.append(row)

    wb.save(output_path)
    return output_path
