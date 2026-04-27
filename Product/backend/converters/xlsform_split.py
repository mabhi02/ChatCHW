"""Aggregate XLSForm + per-module XLSForm workbooks.

The aggregate workbook is the deployable artifact CHT consumes. The
per-module workbooks are review-only: a clinician auditing `mod_fever`
doesn't need to open the full routing graph.

The aggregate workbook is produced by delegating to the existing gen7
converter `backend.converters.json_to_xlsx.convert_to_xlsx`, which already
emits the full CHT-required sheets (survey, choices, settings). Per-module
workbooks are scoped projections of that same content.
"""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import Workbook

logger = logging.getLogger(__name__)


def _as_list(obj) -> list[dict]:
    if isinstance(obj, list):
        return [x for x in obj if isinstance(x, dict)]
    if isinstance(obj, dict):
        return [{**v, "_normalized_id": k} for k, v in obj.items() if isinstance(v, dict)]
    return []


def _id_of(item: dict) -> str:
    return str(item.get("id") or item.get("module_id") or item.get("predicate_id")
               or item.get("_normalized_id") or "")


def build_aggregate_workbook(clinical_logic: dict, out: Path) -> None:
    """Delegate to the existing gen7 xlsx converter.

    The gen7 converter writes a single workbook with survey/choices/settings
    plus any extra sheets; that is exactly the aggregate shape we want.
    """
    from backend.converters.json_to_xlsx import convert_to_xlsx
    convert_to_xlsx(clinical_logic, str(out))


def build_per_module_workbook(module: dict, clinical_logic: dict, out: Path) -> None:
    """Emit a scoped workbook for a single module.

    Sheets:
      - overview     -- module metadata (id, title, priority, trigger/done flags)
      - variables    -- variables this module collects
      - predicates   -- predicates referenced by this module's rules
      - rules        -- the decision table rows
      - routes       -- outgoing routes to other modules
      - phrases      -- phrase_bank entries tagged to this module
    """
    mid = _id_of(module)
    wb = Workbook()

    ws = wb.active
    ws.title = "overview"
    ws.append(["field", "value"])
    for k in ("id", "module_id", "display_name", "title", "priority",
              "trigger_flag", "done_flag", "symptom_trigger"):
        if k in module:
            ws.append([k, str(module.get(k, ""))])
    ws.append(["_manual_page", str(module.get("_manual_page", ""))])
    ws.append(["_order_source", str(module.get("_order_source", ""))])

    # Variables scoped to this module
    ws = wb.create_sheet("variables")
    ws.append(["id", "display_name", "data_type", "unit", "source_quote"])
    module_inputs = set(module.get("inputs") or [])
    for v in _as_list(clinical_logic.get("variables", [])):
        if _id_of(v) in module_inputs:
            ws.append([
                _id_of(v),
                str(v.get("display_name", "")),
                str(v.get("data_type", "")),
                str(v.get("unit", v.get("units", ""))),
                str(v.get("source_quote", ""))[:200],
            ])

    # Predicates referenced by this module's rules
    ws = wb.create_sheet("predicates")
    ws.append(["id", "human_label", "formal_definition", "units", "missingness_rule", "provenance"])
    referenced_preds: set[str] = set()
    for rule in (module.get("rules") or []):
        if isinstance(rule, dict):
            cond = rule.get("condition", "") or ""
            for p in _as_list(clinical_logic.get("predicates", [])):
                if _id_of(p) in cond:
                    referenced_preds.add(_id_of(p))
    for p in _as_list(clinical_logic.get("predicates", [])):
        if _id_of(p) in referenced_preds:
            ws.append([
                _id_of(p),
                str(p.get("human_label", p.get("description_clinical", ""))),
                str(p.get("formal_definition", p.get("threshold_expression", ""))),
                str(p.get("units", "")),
                str(p.get("missingness_rule", "")),
                str(p.get("provenance", "")),
            ])

    # Rules
    ws = wb.create_sheet("rules")
    ws.append(["rule_id", "condition", "outputs_summary", "source_quote"])
    for rule in (module.get("rules") or []):
        if isinstance(rule, dict):
            outputs = rule.get("outputs", {})
            outputs_summary = ", ".join(f"{k}={v}" for k, v in (outputs.items() if isinstance(outputs, dict) else []))
            ws.append([
                str(rule.get("id", rule.get("rule_id", ""))),
                str(rule.get("condition", "")),
                outputs_summary[:400],
                str(rule.get("source_quote", ""))[:200],
            ])

    # Routes out
    ws = wb.create_sheet("routes")
    ws.append(["target_module", "condition"])
    for route in (module.get("routes_to") or []):
        if isinstance(route, dict):
            ws.append([str(route.get("target", "")), str(route.get("condition", ""))])

    # Phrase bank scoped to this module
    ws = wb.create_sheet("phrases")
    ws.append(["id", "category", "text", "module_context"])
    for ph in _as_list(clinical_logic.get("phrase_bank", [])):
        mod_ctx = ph.get("module_context", "")
        if mod_ctx == mid or (isinstance(mod_ctx, list) and mid in mod_ctx):
            ws.append([
                _id_of(ph), str(ph.get("category", "")),
                str(ph.get("text", ""))[:500], str(mod_ctx),
            ])

    wb.save(out)


def emit_all(clinical_logic: dict, output_dir: Path) -> dict:
    """Write aggregate `form.xlsx` plus per-module workbooks. Returns file paths."""
    paths: dict[str, Path] = {}

    aggregate = output_dir / "form.xlsx"
    try:
        build_aggregate_workbook(clinical_logic, aggregate)
        paths["aggregate"] = aggregate
    except Exception as exc:
        logger.warning("xlsform_split: aggregate build failed: %s", exc)

    modules_raw = clinical_logic.get("modules", {})
    modules_iter = modules_raw.items() if isinstance(modules_raw, dict) else (
        (_id_of(m), m) for m in _as_list(modules_raw)
    )
    modules_dir = output_dir / "form_per_module"
    modules_dir.mkdir(exist_ok=True)
    for mid, module in modules_iter:
        if not isinstance(module, dict) or not mid:
            continue
        out = modules_dir / f"form_{mid}.xlsx"
        try:
            build_per_module_workbook(module, clinical_logic, out)
            paths[f"mod_{mid}"] = out
        except Exception as exc:
            logger.warning("xlsform_split: module %s build failed: %s", mid, exc)

    return paths
