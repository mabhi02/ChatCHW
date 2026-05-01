# Normalizes medical-team Excel rule spreadsheets into structured JSON bundles (no execution semantics).
#
# Writes one workbook-level JSON file per source .xlsx into normalized_rules/ for black-box comparator prep.
from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Mapping, MutableMapping, Optional, Sequence, Tuple

import openpyxl

DESCRIPTION_HEADERS_LOWER = {
    "description",
    "purpose / scenario",
    "purpose",
    "question",
    "diagnosis",
    "observation",
    "diagnosis table",
}


def _script_dir() -> Path:
    return Path(__file__).resolve().parent


def _sanitize(parts: Sequence[str]) -> str:
    s = "__".join(p for p in parts if p).lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_") or "unknown_module"


def _coerce_cell(value: Any) -> Any | None:
    if value is None:
        return None
    if isinstance(value, float) and abs(value - int(value)) < 1e-9:
        return int(value)
    if isinstance(value, str):
        t = value.strip()
        return t if t else None
    return value


def _unique_key(d: Mapping[str, Any], key: str) -> str:
    if key not in d:
        return key
    n = 2
    while f"{key}__dup_{n}" in d:
        n += 1
    return f"{key}__dup_{n}"


def _last_used_col(ws: Any, row_idx: int, scan: int = 80) -> int:
    last = 1
    for c in range(1, scan + 1):
        if _coerce_cell(ws.cell(row_idx, c).value) is not None:
            last = c
    return last


def _infer_context(workbook_filename: str, sheet_name: str) -> str:
    s = sheet_name.lower()
    w = workbook_filename.lower()
    if "opening module" in s:
        return "opening_branching"
    if "traffic cop" in s:
        return "traffic_cop_grid"
    if "integrative" in s:
        return "integrative_routing"
    if "dosing" in s:
        return "embedded_dosing_tables"
    if "treatment" in s:
        return "clinical_treatment"
    if "thin module" in s or "swollen module" in s:
        return "nutrition_screen_diagnosis"
    if "diarrhea diagnosis" in s:
        return "diagnosis"
    if s == "sheet1" and "diarrhea" in w:
        return "diagnosis"
    if s.endswith("diagnosis table"):
        return "diagnosis"
    return "clinical_diagnosis_or_mixed"


def _role_for_header(text: str, ctx: str, treatment_dx_as_condition: bool) -> str:
    raw = (text or "").strip()
    if not raw:
        return "ignore"
    lh = raw.lower()
    upper = raw.upper()

    if lh == "#":
        return "meta_rule_id"
    if lh.startswith("#"):
        return "meta_rule_id"
    if upper == "RULE":
        return "meta_rule_id"
    if lh in DESCRIPTION_HEADERS_LOWER:
        return "meta_description"

    if ctx == "traffic_cop_grid":
        if lh in ("start?", "scenario", "branch label"):
            return "meta_route_label"
        if lh.endswith("results") or lh == "results" or lh == "result":
            return "output"
        return "condition"

    if ctx == "opening_branching":
        if lh in ("question", "observation"):
            return "opening_prompt"
        return "output_branch"

    # Treatment tables: dx_* act as preconditions for dosing decisions
    if ctx == "clinical_treatment" and upper.startswith("DX_"):
        return "condition"

    if lh in {"clinical condition", "citation", "action", "results"}:
        return "output"

    prefixes_out = (
        "TX_",
        "RX_",
        "ADV_",
        "REF_HEALTH",
        "NEED_",
        "OUT_",
        "FLAG_",
        "PRIORITY_",
    )
    if any(upper.startswith(p) for p in prefixes_out):
        return "output"

    if upper.startswith("DX_"):
        if ctx in ("diagnosis", "clinical_diagnosis_or_mixed", "nutrition_screen_diagnosis", "integrative_routing"):
            return "output"
        if treatment_dx_as_condition:
            return "condition"
        return "output"

    if "RDT_RESULT" in upper or upper.startswith("LAB_MALARIA_RDT") or upper == "LAB_RDT_MALARIA_RESULT":
        return "output"

    if upper.startswith(("P_DANGER_SIGN", "P_FAST_BREATHING", "P_PRIORITY", "IS_PRIORITY")):
        return "output"

    if lh.startswith("need_"):
        return "output"

    if upper.startswith(("Q_", "EX_", "LAB_MUAC", "DEMO_", "V_", "HX_")):
        return "condition"
    if upper.startswith("LAB_"):
        return "condition"
    if lh.startswith("p_malaria") or lh.startswith("q_") or lh.startswith("chw"):
        return "condition"

    if upper.startswith("TX_") or upper.startswith("RX_"):
        return "output"

    # Default: keep as condition for unknown clinical columns
    return "condition"


def _stop_col_a(val: Any, tokens: Sequence[str]) -> bool:
    if val is None:
        return False
    t = str(val).strip().lower()
    return any(tok in t for tok in tokens)


def _normalize_opening(ws: Any, workbook: str, sheet: str, cfg: Mapping[str, Any]) -> Dict[str, Any]:
    mod = dict(
        module_id=_sanitize([Path(workbook).stem, sheet]),
        source_file=workbook,
        source_sheet=sheet,
        kind="opening_branching",
        sections=[],
    )
    glob = 0
    for sec in cfg["sections"]:
        sec_id = sec["section_id"]
        hdr_row = int(sec["header_row"])
        r0, r1 = int(sec["first_row"]), int(sec["last_row"])
        last_c = _last_used_col(ws, hdr_row)
        headers: List[Tuple[str, int, str]] = []
        for c in range(1, last_c + 1):
            h = _coerce_cell(ws.cell(hdr_row, c).value)
            if h is None:
                continue
            hs = str(h).strip()
            role = _role_for_header(hs, "opening_branching", False)
            headers.append((hs, c, role))

        rules: List[Dict[str, Any]] = []
        ri_local = 0
        for r in range(r0, r1 + 1):
            cond: MutableMapping[str, Any] = {}
            out: MutableMapping[str, Any] = {}
            md: MutableMapping[str, Any] = {"opening_section": sec_id, "opening_row_index": r}
            nonempty = False

            for hs, c, role in headers:
                cell = _coerce_cell(ws.cell(r, c).value)
                if role == "meta_rule_id":
                    nonempty = nonempty or cell is not None
                elif role == "opening_prompt":
                    nonempty = nonempty or cell is not None
                else:
                    if cell is None and role != "opening_prompt":
                        continue
                    if cell is not None:
                        nonempty = True

                if role == "meta_rule_id":
                    nk = hs if hs == "#" else "rule_anchor"
                    md[nk] = cell
                elif role == "meta_description":
                    md["description_cell"] = hs
                    md["description_value"] = cell
                elif role == "opening_prompt":
                    out[_unique_key(out, hs)] = cell
                elif role == "output_branch":
                    out[_unique_key(out, hs)] = cell
                else:
                    nk = _unique_key(out, hs)
                    out[nk] = cell

            if not nonempty:
                continue

            ri_local += 1
            glob += 1
            rules.append(
                dict(
                    rule_index=glob,
                    rule_index_in_section=ri_local,
                    raw_row_index=r,
                    conditions=cond,
                    outputs={k: v for k, v in out.items() if v is not None},
                    notes="",
                    metadata=md,
                )
            )

        mod["sections"].append({"section_id": sec_id, "rules": rules})

    return mod


def _normalize_rule_grid(
    ws: Any,
    workbook: str,
    sheet: str,
    ctx: str,
    header_row: int,
    first_data_row: int,
    stop_tokens: Optional[Sequence[str]] = None,
    module_id_override: Optional[str] = None,
) -> Dict[str, Any]:
    stops = tuple(s.lower() for s in (stop_tokens or ()))
    treatment_dx_as_condition = ctx == "clinical_treatment"
    last_c = _last_used_col(ws, header_row)
    header_plan: List[Tuple[str, int, str]] = []
    for c in range(1, last_c + 1):
        h = _coerce_cell(ws.cell(header_row, c).value)
        if h is None:
            continue
        hs = str(h).strip()
        role = _role_for_header(hs, ctx, treatment_dx_as_condition)
        header_plan.append((hs, c, role))

    rules: List[Dict[str, Any]] = []
    ri = 0
    max_row = ws.max_row or first_data_row
    for r in range(first_data_row, max_row + 1):
        a = ws.cell(r, 1).value
        a_txt = "" if a is None else str(a).strip()
        if stops and _stop_col_a(a, stops):
            break

        nonempty = False
        for _, ci, _ in header_plan:
            if _coerce_cell(ws.cell(r, ci).value) is not None:
                nonempty = True
                break
        if not nonempty:
            continue

        md: MutableMapping[str, Any] = {}
        cond: MutableMapping[str, Any] = {}
        out: MutableMapping[str, Any] = {}

        meta_rule_val = meta_desc_val = None
        meta_desc_hdr = None

        for hs, ci, declared_role in header_plan:
            v = _coerce_cell(ws.cell(r, ci).value)
            role = _role_for_header(hs, ctx, treatment_dx_as_condition)

            if role == "meta_rule_id":
                meta_rule_val = v
                continue
            if role == "meta_description":
                meta_desc_hdr = hs
                meta_desc_val = v
                continue
            if role == "meta_route_label":
                md["route_scenario"] = v
                continue
            if role == "ignore":
                continue

            tgt = cond if role == "condition" else out if role == "output" else None
            if tgt is None:
                continue

            nk = _unique_key(tgt, hs)
            tgt[nk] = v

        cond_f = {k: vv for k, vv in cond.items() if vv is not None}
        out_f = {k: vv for k, vv in out.items() if vv is not None}
        if not cond_f and not out_f:
            continue

        md.update(
            {
                "spreadsheet_anchor_col_a": a_txt or None,
                "spreadsheet_rule_value": meta_rule_val,
                "spreadsheet_description_header": meta_desc_hdr,
                "spreadsheet_description": meta_desc_val,
            }
        )

        ri += 1
        notes_parts = []
        if meta_desc_val is not None:
            notes_parts.append(str(meta_desc_val))

        rules.append(
            dict(
                rule_index=ri,
                raw_row_index=r,
                conditions=cond_f,
                outputs=out_f,
                notes=" — ".join(notes_parts),
                metadata=md,
            )
        )

    return dict(
        module_id=module_id_override or _sanitize([Path(workbook).stem, sheet]),
        source_file=workbook,
        source_sheet=sheet,
        kind="rule_grid",
        sheet_context_detected=ctx,
        rules=rules,
    )


def _normalize_embedded_dosing(ws: Any, workbook: str, sheet: str, cfg: Mapping[str, Any]) -> Dict[str, Any]:
    """Diarrhea dosing-style sheets: nested 'Rule' mini-tables anchored in column B."""
    marker_col = int(cfg["marker_column_index"])
    hdr_start = int(cfg["header_start_column_index"])
    rule_id_col = int(cfg["rule_id_column_index"])
    look_back = int(cfg["context_label_columns"][0]["max_look_back"])

    labels_by_row: Dict[int, str] = {}
    for r in range(1, (ws.max_row or 1) + 1):
        v = _coerce_cell(ws.cell(r, marker_col).value)
        if v is None:
            continue
        s = str(v).strip()
        if ":" in s and len(s) < 100:
            labels_by_row[r] = s

    def nearest_label(rr: int) -> Optional[str]:
        best_row = None
        for r, lbl in labels_by_row.items():
            if r >= rr:
                continue
            d = rr - r
            if d < 1 or d > look_back:
                continue
            if best_row is None or r > best_row:
                best_row = r
        return labels_by_row[best_row] if best_row is not None else None

    rr = 1
    max_row = ws.max_row or 1
    tables: List[Dict[str, Any]] = []

    while rr <= max_row:
        cell_b = ws.cell(rr, marker_col).value
        if cell_b is None or str(cell_b).strip().upper() != cfg["marker_header_cell"].strip().upper():
            rr += 1
            continue

        hdr_row = rr
        last_c = _last_used_col(ws, hdr_row)
        cols: List[Tuple[str, int]] = []
        for c in range(hdr_start, last_c + 1):
            hn = _coerce_cell(ws.cell(hdr_row, c).value)
            if hn is None:
                continue
            cols.append((str(hn).strip(), c))

        rr += 1
        block_rules: List[Dict[str, Any]] = []

        block_idx_start = rr
        li = 0
        while rr <= max_row:
            if _coerce_cell(ws.cell(rr, marker_col).value) and str(ws.cell(rr, marker_col).value).strip().upper() == cfg["marker_header_cell"].strip().upper():
                break
            rule_val = ws.cell(rr, rule_id_col).value
            any_val = False
            for _, cc in cols:
                if _coerce_cell(ws.cell(rr, cc).value) is not None:
                    any_val = True
                    break
            if rule_val is None and not any_val:
                rr += 1
                continue

            md = dict(
                dosing_header_row=hdr_row,
                dosing_section_hint=nearest_label(rr),
                rule_spreadsheet_anchor=_coerce_cell(rule_val),
                block_first_data_row=block_idx_start,
            )
            cond: MutableMapping[str, Any] = {}
            out: MutableMapping[str, Any] = {}
            for hn, cc in cols:
                vv = _coerce_cell(ws.cell(rr, cc).value)
                hl = hn.lower()
                role = "condition" if hl.startswith("age_between") or hl.startswith("demo_") else "output"
                if role == "condition":
                    cond[_unique_key(cond, hn)] = vv
                else:
                    out[_unique_key(out, hn)] = vv

            li += 1
            block_rules.append(
                dict(
                    rule_index=li,
                    raw_row_index=rr,
                    conditions={k: v for k, v in cond.items() if v is not None},
                    outputs={k: v for k, v in out.items() if v is not None},
                    notes="",
                    metadata=md,
                )
            )
            rr += 1

        tables.append(dict(header_row=hdr_row, rules=block_rules))

    merged_rules: List[Dict[str, Any]] = []
    g = 0
    for t_idx, tbl in enumerate(tables, start=1):
        for rule in tbl["rules"]:
            g += 1
            rr = dict(rule)
            rr["rule_index_global"] = g
            rr["metadata"] = dict(rr.get("metadata", {}))
            rr["metadata"]["embedded_table_order"] = t_idx
            merged_rules.append(rr)

    return dict(
        module_id=_sanitize([Path(workbook).stem, sheet]),
        source_file=workbook,
        source_sheet=sheet,
        kind="embedded_dosing_rule_blocks",
        tables_found=len(tables),
        rules=merged_rules,
    )


def normalize_workbook(path: Path, sheet_cfg_global: Mapping[str, Mapping[str, Any]]) -> Dict[str, Any]:
    wb_name = path.name
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    book_cfg = sheet_cfg_global.get(wb_name, {})

    bundle: Dict[str, Any] = dict(
        source_file=wb_name,
        normalized_at_iso=datetime.now(timezone.utc).isoformat(),
        modules=[],
    )

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            spec = dict(book_cfg.get(sheet_name, {}))

            inferred = _infer_context(wb_name, sheet_name)
            mode = spec.get("mode", "infer")

            if mode == "opening_sections":
                bundle["modules"].append(_normalize_opening(ws, wb_name, sheet_name, spec))
                continue

            if mode == "embedded_rule_blocks":
                bundle["modules"].append(_normalize_embedded_dosing(ws, wb_name, sheet_name, spec))
                continue

            if mode == "rule_table":
                hdr_row = int(spec["header_row"])
                first_dr = int(spec.get("first_data_row", hdr_row + 1))
                stops = spec.get("stop_col_a_contains") or ()
                oid = spec.get("module_id")

                ov_ctx = inferred
                m = _normalize_rule_grid(ws, wb_name, sheet_name, ov_ctx, hdr_row, first_dr, stops, module_id_override=oid)
                bundle["modules"].append(m)
                continue

            if mode == "infer":
                hdr_row = int(spec.get("header_row", 1))
                first_dr = int(spec.get("first_data_row", hdr_row + 1))
                stops = spec.get("stop_col_a_contains") or ()
                bundle["modules"].append(
                    _normalize_rule_grid(ws, wb_name, sheet_name, inferred, hdr_row, first_dr, stops)
                )
                continue

    finally:
        wb.close()

    return bundle


def main() -> None:
    base_dir = Path(__file__).resolve().parent
    excel_dir = base_dir.parent
    out_dir = base_dir / "normalized_rules"
    out_dir.mkdir(parents=True, exist_ok=True)
    cfg_root = json.loads((base_dir / "sheet_config.json").read_text(encoding="utf-8"))

    pairs = [
        ("Navigator Modules.xlsx", "navigator.json"),
        ("Diarrhea Module.xlsx", "diarrhea.json"),
        ("Fever (Malaria) Module.xlsx", "fever_malaria.json"),
        ("Pneumonia (Cough) Module.xlsx", "pneumonia_cough.json"),
    ]
    for fn, json_name in pairs:
        src = excel_dir / fn
        if not src.exists():
            raise SystemExit(f"Missing workbook: {src}")
        blob = normalize_workbook(src, cfg_root)
        out_path = out_dir / json_name
        out_path.write_text(json.dumps(blob, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
        print(f"Wrote {out_path} ({len(blob['modules'])} modules)")


if __name__ == "__main__":
    main()
